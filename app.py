import streamlit as st
import pandas as pd
import tempfile
from openpyxl import Workbook

# Streamlit UI
def main():
    st.title("Excel Düzenleme Uygulaması")

    # File upload
    uploaded_file = st.file_uploader("Excel dosyasını yükleyin", type=["xlsx"])

    if uploaded_file:
        try:
            # Load workbook using pandas
            with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp_file:
                tmp_file.write(uploaded_file.read())
                tmp_file.seek(0)
                excel_file = pd.ExcelFile(tmp_file.name)

            # Debug: Display sheet names
            st.write("Sayfa isimleri: ", excel_file.sheet_names)

            # Check required sheets
            required_sheets = ["mağazalar", "ürünler", "Document"]
            missing_sheets = [sheet for sheet in required_sheets if sheet not in excel_file.sheet_names]

            if missing_sheets:
                st.error(f"Eksik sayfalar: {', '.join(missing_sheets)}. Lütfen doğru dosyayı yükleyin.")
                return

            # Load sheets into DataFrames
            stores_df = excel_file.parse("mağazalar")
            products_df = excel_file.parse("ürünler")

            # Create a new workbook for output
            workbook = Workbook()
            document_sheet = workbook.active
            document_sheet.title = "Document"

            # Add headers to the Document sheet
            headers = [
                "StoreTypeCode", "StoreCode", "ItemTypeCode", "ItemCode", 
                "ColorCode", "ItemDim1Code", "ItemDim2Code", "ItemDim3Code", "CoefficientValue"
            ]
            for col_idx, header in enumerate(headers, start=1):
                document_sheet.cell(row=1, column=col_idx, value=header)

            # Populate Document sheet
            row_idx = 2
            for _, store_row in stores_df.iterrows():
                store_code = store_row[0]  # Assuming StoreCode is in the first column
                for _, product_row in products_df.iterrows():
                    item_code = product_row[0]  # Assuming ItemCode is in the first column
                    coefficient_value = product_row[1]  # Assuming CoefficientValue is in the second column

                    document_sheet.cell(row=row_idx, column=1, value=5)  # StoreTypeCode
                    document_sheet.cell(row=row_idx, column=2, value=store_code)  # StoreCode
                    document_sheet.cell(row=row_idx, column=3, value=1)  # ItemTypeCode
                    document_sheet.cell(row=row_idx, column=4, value=item_code)  # ItemCode
                    document_sheet.cell(row=row_idx, column=9, value=coefficient_value)  # CoefficientValue
                    row_idx += 1

            # Save updated workbook with modified name
            original_filename = uploaded_file.name.rsplit(".", 1)[0]  # Extract the original filename without extension
            updated_filename = f"{original_filename}_edited.xlsx"

            output = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
            workbook.save(output.name)

            # Provide download link
            st.success("Düzenleme tamamlandı. Aşağıdaki linkten dosyayı indirebilirsiniz.")
            with open(output.name, "rb") as file:
                st.download_button(
                    label="Düzenlenmiş Excel Dosyasını İndir",
                    data=file,
                    file_name=updated_filename,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
        except Exception as e:
            st.error(f"Bir hata oluştu: {str(e)}")

if __name__ == "__main__":
    main()
